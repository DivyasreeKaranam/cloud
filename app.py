from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import openpyxl
import logging
import os

app = Flask(_name_)
app.secret_key = os.environ.get('SECRET_KEY', 'fallback_secret_key')

attendance_file = "attendance.xlsx"

# Configure logging to output to the console
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(_name_)

def init_attendance_file():
    try:
        wb = openpyxl.load_workbook(attendance_file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Attendance"
        sheet.append(["ID", "Name", "Subject", "Date", "Time"])
        wb.save(attendance_file)
        logger.info(f"Created new attendance file: {attendance_file}")

init_attendance_file()

# Sample users (for demonstration purposes)
users = {"admin": generate_password_hash("password")}

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        if username in users and check_password_hash(users[username], password):
            session["username"] = username
            flash("Logged in successfully!", "success")
            return redirect(url_for("index"))
        else:
            flash("Invalid credentials", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("username", None)
    flash("Logged out successfully!", "success")
    return redirect(url_for("login"))

@app.route("/index")
def index():
    if "username" in session:
        return render_template("index.html")
    else:
        flash("Please log in to access this page", "error")
        return redirect(url_for("login"))

@app.route("/take_attendance", methods=["POST"])
def take_attendance():
    if "username" in session:
        try:
            name = request.form["name"]
            subject = request.form["subject"]
            now = datetime.now()
            date = now.strftime("%Y-%m-%d")
            time = now.strftime("%H:%M:%S")

            wb = openpyxl.load_workbook(attendance_file)
            sheet = wb.active
            row_id = sheet.max_row  # Use row number as unique identifier
            sheet.append([row_id, name, subject, date, time])
            wb.save(attendance_file)

            flash("Attendance recorded successfully!", "success")
            return redirect(url_for("show_attendance"))
        except Exception as e:
            logger.error(f"Error recording attendance: {str(e)}")
            flash("An error occurred while recording attendance", "error")
            return redirect(url_for("index"))
    else:
        flash("Please log in to take attendance", "error")
        return redirect(url_for("login"))

@app.route("/attendance")
def show_attendance():
    if "username" in session:
        try:
            wb = openpyxl.load_workbook(attendance_file)
            sheet = wb.active
            attendance = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                id, name, subject, date, time = row
                attendance[id] = (f"{name} ({subject})", f"{date} {time}")
            return render_template("attendance.html", attendance=attendance)
        except Exception as e:
            logger.error(f"Error displaying attendance: {str(e)}")
            flash("An error occurred while fetching attendance data", "error")
            return redirect(url_for("index"))
    else:
        flash("Please log in to view attendance", "error")
        return redirect(url_for("login"))

@app.route("/delete_attendance/<int:id>", methods=["POST"])
def delete_attendance(id):
    if "username" in session:
        try:
            wb = openpyxl.load_workbook(attendance_file)
            sheet = wb.active
            
            # Find the row with the matching ID and delete it
            for row in range(2, sheet.max_row + 1):  # Start from 2 to skip header
                if sheet.cell(row=row, column=1).value == id:
                    sheet.delete_rows(row, 1)
                    break
            
            wb.save(attendance_file)
            flash("Attendance record deleted successfully!", "success")
        except Exception as e:
            logger.error(f"Error deleting attendance: {str(e)}")
            flash("An error occurred while deleting the attendance record", "error")
        
        return redirect(url_for("show_attendance"))
    else:
        flash("Please log in to delete attendance records", "error")
        return redirect(url_for("login"))

if _name_ == "_main_":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
